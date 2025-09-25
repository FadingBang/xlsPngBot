'''
This File contains the code to convert an image to an excel spreadsheet,

'''
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill
import openpyxl
from PIL import Image
import gspread
from oauth2client.service_account import ServiceAccountCredentials
import re

'''
This function will take in an image name/path, read in the image's rgb data for it's pixels, and then write these pixels to an excel spreadsheet,
To create this effect is will read the rgb data, and then color 3 cells in the spreadsheet to match. So there will be 1 red cell, 1 green cell and 1 blue cell
in the spreadsheet for every pixel. The magnitude of the individual colors will reflect the intensity of the rgb data and when viewed at a good scale will make the spreadsheet
look like the image. 


it has 2 optional arguments, resolution controls how detailed the excel image will be, a value of 1 means for every row/column there will be a dedicated row/column
a value less than 1 means that it will skip over some pixels (i.e. resolution = .5 means there will be only half as many 'pixels' in the resulting spreadsheet)
'''
def Convert_Image_To_Excel_Spreadsheet(image_path, resolution=.3, output_file="results.xlsx"):
    wb = openpyxl.Workbook() #first create an excel workbook
    sheet = wb.get_sheet_by_name("Sheet") #each workbook has a default sheet, so we grab it to use
    im = Image.open(image_path)#next we read in our image
    width, height = im.size #get the width and height of the image, this will correspond to how many pixels tall and wide the image is.
    rgb_im = im.convert('RGB') #now get the rgb data for the image.

    #now we need to loop through the image, taking into account the resolution,
    for x in range(0, int(width * resolution)):
        #for each pixel we use 3 cells in the spreadsheet, each in the same row,
        #so only the column value will change for each cell in the pixel.
        #also important to note, excel is base 1, so a value of 0 is invalid, this is a slight diversion from how
        #we normally code stuff...so we need to take that into acctoun

        #calculate and store the colum_indicies for this column.
        column_indicies = [get_column_letter((x * 3) + 1), get_column_letter((x * 3) + 2), get_column_letter((x * 3) + 3)]

        #in order to preserve the aspect ratio of the image we want to scale down the width of the pixels.
        #Our goal is to get 3 cells in a neat square, according to the documentation the height of a cell is 10 px,
        #since we have 3 this would normally mean we would have the width be 10/3, however this still resulted in a more rectangle shape,
        #with a little tweaking the below creates a close enough square for our 'pixels'
        sheet.column_dimensions[column_indicies[0]].width = 10 / 9
        sheet.column_dimensions[column_indicies[1]].width = 10 / 9
        sheet.column_dimensions[column_indicies[2]].width = 10 / 9

        #next we loop through all of the rows for this set of columns.
        for row in range(1, int(height * resolution)):
            #now we get the rgb data for this pixel. we store this in rgb_data
            rgb_data = rgb_im.getpixel((int(x / resolution), int(row / resolution)))
            for i in range(3):#next we loop through the 3 cells for this pixel.
                colors = [0, 0, 0]#start with pure white,
                colors[i] = rgb_data[i]#update a single value from our rgb data
                col = get_column_letter((x * 3) + i + 1)#since excel uses a A1 type of  coordinate system we need to convert the column number with a letter, thankfully there is allready a function for that in our libraries,

                #now get the cell
                cell = sheet[col + str(row)]

                #just for propserity, I also store the rgb value for this cell, just so it's not blank,
                cell.value = rgb_data[i]

                #next we recolor the cell.
                #the libraries use a string based color scheme, this code takes our colors list and combines them in a way that the libraries
                #can make sense of and convert to a color for the cell.
                color_string = "".join([str(hex(i))[2:].upper().rjust(2, "0") for i in colors])
                cell.fill = PatternFill(fill_type="solid", start_color='FF' + color_string,
                                        end_color='FF' + color_string)
    #don't forget to save.
    wb.save(output_file)

def Convert_Google_Sheet_To_Image(spreadsheet_url, output_file="output_image.png", credentials_file="credentials.json"):
    """
    This function converts a Google spreadsheet to an image.
    It reads the RGB values from the cells and reconstructs the image.
    
    Args:
        spreadsheet_url (str): URL of the Google Spreadsheet
        output_file (str): Path where to save the output image (default: "output_image.png")
        credentials_file (str): Path to the Google Sheets API credentials file
    """
    # Setup the credentials
    scope = ['https://spreadsheets.google.com/feeds', 'https://www.googleapis.com/auth/drive']
    credentials = ServiceAccountCredentials.from_json_keyfile_name(credentials_file, scope)
    client = gspread.authorize(credentials)
    
    # Extract spreadsheet ID from URL
    match = re.search(r'/d/([a-zA-Z0-9-_]+)', spreadsheet_url)
    if match:
        spreadsheet_id = match.group(1)
    else:
        raise ValueError("Invalid Google Sheets URL")
    
    # Open the spreadsheet
    sheet = client.open_by_key(spreadsheet_id).sheet1

    # Find the dimensions of the image
    max_row = sheet.max_row
    max_col = sheet.max_column
    width = max_col // 3  # Since we used 3 cells per pixel
    height = max_row

    # Create a new image with the same dimensions
    image = Image.new('RGB', (width, height))
    pixels = []

    # Read the RGB values from the Excel cells
    for row in range(1, max_row + 1):
        row_pixels = []
        for x in range(width):
            # Get RGB values from the three cells representing one pixel
            r = int(sheet[f"{get_column_letter(x * 3 + 1)}{row}"].value or 0)
            g = int(sheet[f"{get_column_letter(x * 3 + 2)}{row}"].value or 0)
            b = int(sheet[f"{get_column_letter(x * 3 + 3)}{row}"].value or 0)
            row_pixels.append((r, g, b))
        pixels.append(row_pixels)

    # Put the pixels in the image
    for y in range(height):
        for x in range(width):
            if y < len(pixels) and x < len(pixels[y]):
                image.putpixel((x, y), pixels[y][x])

    # Save the image
    image.save(output_file)